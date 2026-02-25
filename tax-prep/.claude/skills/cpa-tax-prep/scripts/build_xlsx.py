#!/usr/bin/env python3
"""
Deterministic XLSX builder for CPA tax packages.

Reads a structured JSON file and produces an 8-tab xlsx workbook.
All formatting decisions are made here — the JSON contains raw data only.

Usage:
    python build_xlsx.py <json_file> <output_xlsx>
"""

import json
import os
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Shared Styles ────────────────────────────────────────────────────────────

BOLD = Font(bold=True)
BOLD_LARGE = Font(bold=True, size=14)
BOLD_SECTION = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
CURRENCY_FMT = '#,##0.00'
PCT_FMT = '0.00%'


# ── Helpers ──────────────────────────────────────────────────────────────────

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')


def currency_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = CURRENCY_FMT
    return cell


def bold_currency_cell(ws, row, col, value):
    cell = currency_cell(ws, row, col, value)
    cell.font = BOLD
    return cell


def total_fill_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).fill = TOTAL_FILL


# ── Tab Builders ─────────────────────────────────────────────────────────────

def build_summary_tab(wb, data):
    """Build the Summary tab — filer info, annual totals (cross-sheet refs), CPA notes."""
    ws = wb.active
    ws.title = "Summary"
    set_col_widths(ws, [30, 50])

    year = data["meta"]["year"]
    filer = data["summary"]["filer"]

    ws.cell(row=1, column=1,
            value=f"{year} Tax Preparation — {filer['name']} / {filer['business']}").font = BOLD_LARGE

    # Filer info
    ws.cell(row=3, column=1, value="FILER INFO").font = BOLD_SECTION

    filer_fields = [
        ("Name", filer.get("name", "")),
        ("Business", filer.get("business", "")),
        ("Entity", filer.get("entity", "")),
        ("State(s)", filer.get("states", "")),
    ]
    if filer.get("filing_status"):
        filer_fields.append(("Filing Status", filer["filing_status"]))
    filer_fields.append(("Address", filer.get("address", "")))
    if filer.get("address_change"):
        filer_fields.append(("Address Change", filer["address_change"]))
    if filer.get("phone"):
        filer_fields.append(("Phone", filer["phone"]))
    if filer.get("email"):
        filer_fields.append(("Email", filer["email"]))

    for i, (label, value) in enumerate(filer_fields):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=2, value=value)

    # Annual totals — use cross-sheet formula references
    totals_start = 4 + len(filer_fields) + 1
    ws.cell(row=totals_start, column=1, value="ANNUAL TOTALS").font = BOLD_SECTION

    # We'll build each total row and track where the formulas point.
    # The exact cell references depend on what the other tab builders produce,
    # so we store marker names and resolve them after all tabs are built.
    # For now, use placeholder rows — we'll fill in formulas in main().
    totals_labels = [
        "Gross Business Income",
        "Interest Income",
        "Business Expenses",
        "Home Office Deduction",
        "Health Insurance",
        "Est. Taxes (Federal)",
        "Est. Taxes (State)",
        "Retirement",
    ]

    for i, label in enumerate(totals_labels):
        r = totals_start + 1 + i
        ws.cell(row=r, column=1, value=label).font = BOLD

    # CPA questions
    questions_start = totals_start + 1 + len(totals_labels) + 1
    questions = data["summary"].get("cpa_questions", [])
    ws.cell(row=questions_start, column=1, value="QUESTIONS FOR CPA").font = BOLD_SECTION
    for i, question in enumerate(questions):
        ws.cell(row=questions_start + 1 + i, column=1, value=question)

    # CPA notes
    notes_start = questions_start + 1 + len(questions) + 1
    notes = data["summary"].get("cpa_notes", [])
    ws.cell(row=notes_start, column=1, value="NOTES").font = BOLD_SECTION
    for i, note in enumerate(notes):
        ws.cell(row=notes_start + 1 + i, column=1, value=note)

    # Return layout info for cross-sheet formula wiring
    return {
        "totals_start_row": totals_start + 1,
        "labels": totals_labels,
    }


def build_income_tab(wb, data):
    """Build the Income tab — monthly breakdown, interest, by-source."""
    ws = wb.create_sheet("Income")
    set_col_widths(ws, [15, 18, 70])

    year = data["meta"]["year"]
    filer = data["summary"]["filer"]
    income = data["income"]

    ws.cell(row=1, column=1, value=f"{year} Income — {filer['business']}").font = BOLD_LARGE

    # Monthly income
    header_row(ws, 3, ["Month", "Amount", "Sources"])

    monthly = income.get("monthly", [])
    for i, item in enumerate(monthly):
        r = 4 + i
        ws.cell(row=r, column=1, value=item["month"])
        currency_cell(ws, r, 2, item["amount"])
        ws.cell(row=r, column=3, value=item.get("sources", ""))

    total_row = 4 + len(monthly)
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD
    c = ws.cell(row=total_row, column=2)
    c.value = f"=SUM(B4:B{total_row - 1})"
    c.number_format = CURRENCY_FMT
    c.font = BOLD
    total_fill_row(ws, total_row, 3)

    # Interest income
    interest_start = total_row + 2
    interest_items = income.get("interest", [])
    if interest_items:
        ws.cell(row=interest_start, column=1, value="Interest Income").font = BOLD_SECTION
        for i, item in enumerate(interest_items):
            r = interest_start + 1 + i
            ws.cell(row=r, column=1, value=item["source"])
            currency_cell(ws, r, 2, item["amount"])
        interest_total_row = interest_start + 1 + len(interest_items)
    else:
        interest_total_row = interest_start

    # Income by source
    source_start = interest_total_row + 1
    by_source = income.get("by_source", [])
    if by_source:
        ws.cell(row=source_start, column=1, value="Income by Source").font = BOLD_SECTION
        header_row(ws, source_start + 1, ["Source", "Annual Total", ""])
        for i, item in enumerate(by_source):
            r = source_start + 2 + i
            ws.cell(row=r, column=1, value=item["source"])
            currency_cell(ws, r, 2, item["annual_total"])

    # Return refs for cross-sheet formulas
    return {
        "income_total_cell": f"B{total_row}",
        "interest_items": interest_items,
        "interest_start_row": interest_start + 1 if interest_items else None,
    }


def build_expenses_tab(wb, data):
    """Build the Business Expenses tab — recurring, one-time, excluded."""
    ws = wb.create_sheet("Business Expenses")
    set_col_widths(ws, [30, 18, 50, 18, 35, 40])

    year = data["meta"]["year"]
    filer = data["summary"]["filer"]
    expenses = data["business_expenses"]

    ws.cell(row=1, column=1, value=f"{year} Business Expenses — {filer['business']}").font = BOLD_LARGE

    # Section 1: Monthly Recurring
    ws.cell(row=3, column=1, value="MONTHLY RECURRING").font = BOLD_SECTION
    header_row(ws, 4, ["Vendor", "Monthly Amount", "Months Charged", "Annual Total", "Category", ""])

    recurring = expenses.get("recurring", [])
    for i, item in enumerate(recurring):
        r = 5 + i
        ws.cell(row=r, column=1, value=item["vendor"])
        monthly_amt = item["monthly_amount"]
        if isinstance(monthly_amt, (int, float)):
            currency_cell(ws, r, 2, monthly_amt)
        else:
            ws.cell(row=r, column=2, value=monthly_amt)
        ws.cell(row=r, column=3, value=item["months_charged"])
        currency_cell(ws, r, 4, item["annual_total"])
        ws.cell(row=r, column=5, value=item["category"])

    recurring_end = 5 + len(recurring) - 1 if recurring else 5
    recurring_total_row = recurring_end + 1 if recurring else 6
    ws.cell(row=recurring_total_row, column=1, value="Recurring Subtotal").font = BOLD
    c = ws.cell(row=recurring_total_row, column=4)
    if recurring:
        c.value = f"=SUM(D5:D{recurring_end})"
    else:
        c.value = 0
    c.number_format = CURRENCY_FMT
    c.font = BOLD
    total_fill_row(ws, recurring_total_row, 5)

    # Section 2: One-Time / Non-Recurring
    onetime_section_row = recurring_total_row + 2
    ws.cell(row=onetime_section_row, column=1, value="ONE-TIME / NON-RECURRING").font = BOLD_SECTION
    header_row(ws, onetime_section_row + 1, ["Date", "Vendor", "Amount", "Category", "Description", ""])

    one_time = expenses.get("one_time", [])
    onetime_data_start = onetime_section_row + 2
    for i, item in enumerate(one_time):
        r = onetime_data_start + i
        ws.cell(row=r, column=1, value=item["date"])
        ws.cell(row=r, column=2, value=item["vendor"])
        currency_cell(ws, r, 3, item["amount"])
        ws.cell(row=r, column=4, value=item["category"])
        ws.cell(row=r, column=5, value=item.get("description", ""))

    onetime_end = onetime_data_start + len(one_time) - 1 if one_time else onetime_data_start
    onetime_total_row = onetime_end + 1 if one_time else onetime_data_start + 1
    ws.cell(row=onetime_total_row, column=1, value="One-Time Subtotal").font = BOLD
    c = ws.cell(row=onetime_total_row, column=3)
    if one_time:
        c.value = f"=SUM(C{onetime_data_start}:C{onetime_end})"
    else:
        c.value = 0
    c.number_format = CURRENCY_FMT
    c.font = BOLD
    total_fill_row(ws, onetime_total_row, 5)

    # Grand total
    grand_row = onetime_total_row + 2
    ws.cell(row=grand_row, column=1, value="GRAND TOTAL — ALL BUSINESS EXPENSES").font = BOLD
    c = ws.cell(row=grand_row, column=3)
    c.value = f"=D{recurring_total_row}+C{onetime_total_row}"
    c.number_format = CURRENCY_FMT
    c.font = BOLD
    c2 = ws.cell(row=grand_row, column=4)
    c2.value = f"=D{recurring_total_row}+C{onetime_total_row}"
    c2.number_format = CURRENCY_FMT
    c2.font = BOLD
    total_fill_row(ws, grand_row, 5)

    # Excluded items
    excluded = expenses.get("excluded", [])
    if excluded:
        note_row = grand_row + 2
        ws.cell(row=note_row, column=1,
                value="EXCLUDED FROM EXPENSES (Owner Draws / Prior-Year Taxes):").font = BOLD
        header_row(ws, note_row + 1, ["Date", "Payee", "Amount", "Notes", "", ""])
        for i, item in enumerate(excluded):
            r = note_row + 2 + i
            ws.cell(row=r, column=1, value=item["date"])
            ws.cell(row=r, column=2, value=item["payee"])
            currency_cell(ws, r, 3, item["amount"])
            ws.cell(row=r, column=4, value=item.get("notes", ""))

    return {
        "grand_total_cell": f"D{grand_row}",
    }


def build_personal_deductions_tab(wb, data):
    """Build the Personal Deductions tab."""
    ws = wb.create_sheet("Personal Deductions")
    set_col_widths(ws, [20, 35, 18, 18, 20, 50])

    year = data["meta"]["year"]

    ws.cell(row=1, column=1, value=f"{year} Personal Deductions").font = BOLD_LARGE
    header_row(ws, 3, ["Type", "Vendor/Payee", "Monthly Amount", "Annual Total", "Payment Method", "Notes"])

    items = data.get("personal_deductions", [])
    for i, item in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=item["type"])
        ws.cell(row=r, column=2, value=item["vendor"])
        monthly = item.get("monthly_amount")
        if monthly is not None:
            currency_cell(ws, r, 3, monthly)
        currency_cell(ws, r, 4, item["annual_total"])
        ws.cell(row=r, column=5, value=item.get("payment_method", ""))
        ws.cell(row=r, column=6, value=item.get("notes", ""))

    total_r = 4 + len(items) if items else 5
    ws.cell(row=total_r, column=1, value="TOTAL").font = BOLD
    c = ws.cell(row=total_r, column=4)
    if items:
        c.value = f"=SUM(D4:D{total_r - 1})"
    else:
        c.value = 0
    c.number_format = CURRENCY_FMT
    c.font = BOLD
    total_fill_row(ws, total_r, 6)

    return {}


def build_home_office_tab(wb, data):
    """Build the Home Office tab — row-based per location, with combined total."""
    ws = wb.create_sheet("Home Office")
    set_col_widths(ws, [25, 40])

    year = data["meta"]["year"]
    home_office = data["home_office"]
    locations = home_office.get("locations", [])

    ws.cell(row=1, column=1, value=f"HOME OFFICE DEDUCTION — {year}").font = BOLD_LARGE

    r = 3
    subtotal_rows = []

    for loc_idx, loc in enumerate(locations):
        ws.cell(row=r, column=1,
                value=f"LOCATION {loc_idx + 1}: {loc['description']}").font = BOLD_SECTION

        fields = [
            ("Months", loc["months"]),
            ("Office Sq Ft", loc["office_sqft"]),
            ("Home Sq Ft", loc["home_sqft"]),
            ("Office %", f"{loc['office_pct']}%"),
            ("Monthly Rent", None),  # currency
            ("Monthly Deduction", None),  # currency
            (f"Subtotal ({loc['months_count']} months)", None),  # currency
        ]

        for i, (label, value) in enumerate(fields):
            row = r + 1 + i
            ws.cell(row=row, column=1, value=label).font = BOLD
            if label == "Monthly Rent":
                currency_cell(ws, row, 2, loc["monthly_rent"])
            elif label == "Monthly Deduction":
                currency_cell(ws, row, 2, loc["monthly_deduction"])
            elif label.startswith("Subtotal"):
                bold_currency_cell(ws, row, 2, loc["subtotal"])
                subtotal_rows.append(row)
            else:
                ws.cell(row=row, column=2, value=value)

        r = r + 1 + len(fields) + 1  # gap before next location

    # Total deduction
    total_row = r
    ws.cell(row=total_row, column=1, value="TOTAL HOME OFFICE DEDUCTION").font = BOLD
    if len(subtotal_rows) == 1:
        bold_currency_cell(ws, total_row, 2, locations[0]["subtotal"])
    elif len(subtotal_rows) > 1:
        refs = "+".join(f"B{sr}" for sr in subtotal_rows)
        c = ws.cell(row=total_row, column=2, value=f"={refs}")
        c.number_format = CURRENCY_FMT
        c.font = BOLD
    else:
        bold_currency_cell(ws, total_row, 2, 0)
    total_fill_row(ws, total_row, 2)

    # Notes
    notes = home_office.get("notes", [])
    if notes:
        note_row = total_row + 2
        ws.cell(row=note_row, column=1, value="Note:").font = BOLD
        for i, note in enumerate(notes):
            ws.cell(row=note_row + i, column=2, value=note)

    return {
        "total_cell": f"B{total_row}",
    }


def build_estimated_taxes_tab(wb, data):
    """Build the Estimated Taxes tab — payments, subtotals, prior-year."""
    ws = wb.create_sheet("Estimated Taxes")
    set_col_widths(ws, [15, 12, 20, 18, 30])

    year = data["meta"]["year"]
    est = data["estimated_taxes"]

    ws.cell(row=1, column=1, value=f"{year} Estimated Tax Payments").font = BOLD_LARGE
    header_row(ws, 3, ["Authority", "Quarter", "Date Paid", "Amount", "Confirmation #"])

    payments = est.get("payments", [])

    # Group payments by authority for subtotals
    authorities = {}
    for p in payments:
        auth = p["authority"]
        if auth not in authorities:
            authorities[auth] = []
        authorities[auth].append(p)

    r = 4
    payment_start = r
    for item in payments:
        ws.cell(row=r, column=1, value=item["authority"])
        ws.cell(row=r, column=2, value=item["quarter"])
        ws.cell(row=r, column=3, value=item["date_paid"])
        currency_cell(ws, r, 4, item["amount"])
        ws.cell(row=r, column=5, value=item.get("confirmation", ""))
        r += 1

    payment_end = r - 1

    # Subtotals per authority
    r += 1  # blank row
    subtotal_cells = {}
    federal_total_row = None
    state_total_row = None

    for auth, items in authorities.items():
        ws.cell(row=r, column=1, value=f"{auth} Total").font = BOLD
        # Find the rows for this authority
        auth_rows = []
        for i, p in enumerate(payments):
            if p["authority"] == auth:
                auth_rows.append(payment_start + i)
        if auth_rows:
            refs = ",".join(f"D{ar}" for ar in auth_rows)
            c = ws.cell(row=r, column=4)
            c.value = f"=SUM({refs})"
            c.number_format = CURRENCY_FMT
            c.font = BOLD
        total_fill_row(ws, r, 5)
        subtotal_cells[auth] = f"D{r}"

        if auth == "IRS":
            federal_total_row = r
        else:
            state_total_row = r
        r += 1

    # If no state payments, add a zero line
    if state_total_row is None:
        ws.cell(row=r, column=1, value="State Total").font = BOLD
        currency_cell(ws, r, 4, 0.00)
        ws.cell(row=r, column=5, value="No state income tax" if not any(
            a != "IRS" for a in authorities) else "")
        state_total_row = r
        r += 1

    # Grand total
    r += 1
    ws.cell(row=r, column=1, value="GRAND TOTAL").font = BOLD
    all_subtotal_refs = "+".join(subtotal_cells.values())
    if all_subtotal_refs:
        c = ws.cell(row=r, column=4)
        c.value = f"={all_subtotal_refs}"
        c.number_format = CURRENCY_FMT
        c.font = BOLD
    else:
        bold_currency_cell(ws, r, 4, 0)
    total_fill_row(ws, r, 5)

    # Notes
    notes = est.get("notes", [])
    if notes:
        r += 2
        ws.cell(row=r, column=1, value="Notes:").font = BOLD
        for i, note in enumerate(notes):
            ws.cell(row=r + 1 + i, column=1, value=note)
        r += 1 + len(notes)

    # Prior-year payments
    prior = est.get("prior_year_payments", [])
    if prior:
        r += 1
        ws.cell(row=r, column=1,
                value="PRIOR-YEAR PAYMENTS (from business account — ASK CPA):").font = BOLD_SECTION
        r += 1
        header_row(ws, r, ["Authority", "", "Date Paid", "Amount", "Notes"])
        for i, item in enumerate(prior):
            pr = r + 1 + i
            ws.cell(row=pr, column=1, value=item["authority"])
            ws.cell(row=pr, column=3, value=item["date_paid"])
            currency_cell(ws, pr, 4, item["amount"])
            ws.cell(row=pr, column=5, value=item.get("notes", ""))

    return {
        "federal_total_cell": f"D{federal_total_row}" if federal_total_row else None,
        "state_total_cell": f"D{state_total_row}" if state_total_row else None,
    }


def build_health_insurance_tab(wb, data):
    """Build the Health Insurance tab."""
    ws = wb.create_sheet("Health Insurance")
    set_col_widths(ws, [25, 18, 15, 15, 12, 18, 25])

    year = data["meta"]["year"]

    ws.cell(row=1, column=1, value=f"{year} Health Insurance").font = BOLD_LARGE
    header_row(ws, 3, ["Provider", "Monthly Premium", "Start Month", "End Month",
                        "Months", "Annual Total", "Payment Method"])

    items = data.get("health_insurance", [])
    for i, item in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=1, value=item["provider"])
        currency_cell(ws, r, 2, item["monthly_premium"])
        ws.cell(row=r, column=3, value=item["start_month"])
        ws.cell(row=r, column=4, value=item["end_month"])
        ws.cell(row=r, column=5, value=item["months"])
        # Use formula: premium * months
        c = ws.cell(row=r, column=6)
        c.value = f"=B{r}*E{r}"
        c.number_format = CURRENCY_FMT
        ws.cell(row=r, column=7, value=item.get("payment_method", ""))

    # Total row
    total_row = 4 + len(items) + 1 if items else 6
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD
    c = ws.cell(row=total_row, column=6)
    if items:
        c.value = f"=SUM(F4:F{4 + len(items) - 1})"
    else:
        c.value = 0
    c.number_format = CURRENCY_FMT
    c.font = BOLD
    total_fill_row(ws, total_row, 7)

    # Notes — collect from all insurance items
    all_notes = []
    for item in items:
        all_notes.extend(item.get("notes", []))
    if all_notes:
        note_row = total_row + 2
        ws.cell(row=note_row, column=1, value="Notes:").font = BOLD
        for i, note in enumerate(all_notes):
            ws.cell(row=note_row + 1 + i, column=1, value=note)

    return {
        "total_cell": f"F{total_row}",
    }


def build_retirement_tab(wb, data):
    """Build the Retirement tab."""
    ws = wb.create_sheet("Retirement")
    set_col_widths(ws, [20, 25, 18, 20, 15])

    year = data["meta"]["year"]
    retirement = data["retirement"]

    ws.cell(row=1, column=1, value=f"{year} Retirement Contributions").font = BOLD_LARGE
    header_row(ws, 3, ["Account Type", "Institution", "Amount", "Date", "Tax Year"])

    contributions = retirement.get("contributions", [])
    if contributions:
        for i, item in enumerate(contributions):
            r = 4 + i
            ws.cell(row=r, column=1, value=item["account_type"])
            ws.cell(row=r, column=2, value=item["institution"])
            currency_cell(ws, r, 3, item["amount"])
            ws.cell(row=r, column=4, value=item["date"])
            ws.cell(row=r, column=5, value=item["tax_year"])

        total_row = 4 + len(contributions)
        ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD
        c = ws.cell(row=total_row, column=3)
        c.value = f"=SUM(C4:C{total_row - 1})"
        c.number_format = CURRENCY_FMT
        c.font = BOLD
        total_fill_row(ws, total_row, 5)
        notes_start = total_row + 2
        retirement_total_cell = f"C{total_row}"
    else:
        ws.cell(row=4, column=1, value="(No contributions made yet)")
        notes_start = 6
        retirement_total_cell = None

    # Notes
    notes = retirement.get("notes", [])
    if notes:
        ws.cell(row=notes_start, column=1, value="Notes:").font = BOLD
        for i, note in enumerate(notes):
            ws.cell(row=notes_start + 1 + i, column=1, value=note)

    return {
        "total_cell": retirement_total_cell,
    }


# ── Cross-sheet formula wiring ───────────────────────────────────────────────

def wire_summary_formulas(wb, summary_info, income_info, expenses_info,
                          home_office_info, health_info, est_taxes_info,
                          retirement_info, data):
    """Fill in Summary tab annual totals with cross-sheet formula references."""
    ws = wb["Summary"]
    start = summary_info["totals_start_row"]
    labels = summary_info["labels"]

    for i, label in enumerate(labels):
        r = start + i
        cell = ws.cell(row=r, column=2)

        if label == "Gross Business Income" and income_info.get("income_total_cell"):
            cell.value = f"='Income'!{income_info['income_total_cell']}"
        elif label == "Interest Income":
            interest = data["income"].get("interest", [])
            total = sum(item["amount"] for item in interest)
            cell.value = total
        elif label == "Business Expenses" and expenses_info.get("grand_total_cell"):
            cell.value = f"='Business Expenses'!{expenses_info['grand_total_cell']}"
        elif label == "Home Office Deduction" and home_office_info.get("total_cell"):
            cell.value = f"='Home Office'!{home_office_info['total_cell']}"
        elif label == "Health Insurance" and health_info.get("total_cell"):
            cell.value = f"='Health Insurance'!{health_info['total_cell']}"
        elif label == "Est. Taxes (Federal)" and est_taxes_info.get("federal_total_cell"):
            cell.value = f"='Estimated Taxes'!{est_taxes_info['federal_total_cell']}"
        elif label == "Est. Taxes (State)" and est_taxes_info.get("state_total_cell"):
            cell.value = f"='Estimated Taxes'!{est_taxes_info['state_total_cell']}"
        elif label == "Retirement" and retirement_info.get("total_cell"):
            cell.value = f"='Retirement'!{retirement_info['total_cell']}"
        elif label == "Retirement":
            cell.value = 0
        elif label == "Est. Taxes (State)":
            cell.value = 0

        cell.number_format = CURRENCY_FMT


# ── Main ─────────────────────────────────────────────────────────────────────

def run_recalc(xlsx_path):
    """Run recalc.py on the generated xlsx if LibreOffice is available."""
    recalc_script = os.path.join(
        os.path.dirname(__file__), "..", "..", "xlsx", "scripts", "recalc.py"
    )
    recalc_script = os.path.normpath(recalc_script)

    if not os.path.exists(recalc_script):
        return {"skipped": True, "reason": "recalc.py not found"}

    try:
        abs_xlsx = os.path.abspath(xlsx_path)
        result = subprocess.run(
            [sys.executable, recalc_script, abs_xlsx],
            capture_output=True, text=True, timeout=60,
            cwd=os.path.dirname(recalc_script)
        )
        if result.returncode == 0:
            return json.loads(result.stdout)
        else:
            return {"error": result.stderr or "recalc failed"}
    except Exception as e:
        return {"skipped": True, "reason": str(e)}


def main(json_path, output_path):
    with open(json_path, "r") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # Build each tab
    summary_info = build_summary_tab(wb, data)
    income_info = build_income_tab(wb, data)
    expenses_info = build_expenses_tab(wb, data)
    build_personal_deductions_tab(wb, data)
    home_office_info = build_home_office_tab(wb, data)
    est_taxes_info = build_estimated_taxes_tab(wb, data)
    health_info = build_health_insurance_tab(wb, data)
    retirement_info = build_retirement_tab(wb, data)

    # Wire cross-sheet formulas in Summary
    wire_summary_formulas(
        wb, summary_info, income_info, expenses_info,
        home_office_info, health_info, est_taxes_info,
        retirement_info, data
    )

    # Save
    wb.save(output_path)

    # Run recalc
    recalc_result = run_recalc(output_path)

    # Compute totals from JSON for the summary output
    income_total = sum(m["amount"] for m in data["income"].get("monthly", []))
    interest_total = sum(i["amount"] for i in data["income"].get("interest", []))
    recurring_total = sum(r["annual_total"] for r in data["business_expenses"].get("recurring", []))
    onetime_total = sum(o["amount"] for o in data["business_expenses"].get("one_time", []))
    expense_total = recurring_total + onetime_total
    home_office_total = sum(l["subtotal"] for l in data["home_office"].get("locations", []))
    health_total = sum(h["annual_total"] for h in data.get("health_insurance", []))
    est_tax_total = sum(p["amount"] for p in data["estimated_taxes"].get("payments", []))
    retirement_total = sum(c["amount"] for c in data["retirement"].get("contributions", []))

    result = {
        "status": "success",
        "output": output_path,
        "tabs": 8,
        "totals": {
            "income": income_total,
            "interest": interest_total,
            "business_expenses": expense_total,
            "home_office": home_office_total,
            "health_insurance": health_total,
            "estimated_taxes": est_tax_total,
            "retirement": retirement_total,
        },
        "recalc": recalc_result,
    }

    print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python build_xlsx.py <json_file> <output_xlsx>")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
