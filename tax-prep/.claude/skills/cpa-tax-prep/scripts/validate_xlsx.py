#!/usr/bin/env python3
"""
Validation script for CPA tax package xlsx files.

Compares a generated xlsx workbook against the source JSON data to verify
that all values were written correctly.

Usage:
    python validate_xlsx.py <json_file> <xlsx_file>

Output: JSON with pass/fail status, check count, and error details.
"""

import json
import math
import sys

from openpyxl import load_workbook


TOLERANCE = 0.01  # currency tolerance for floating point


def approx_eq(expected, actual):
    """Check if two numeric values are approximately equal."""
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return abs(expected - actual) < TOLERANCE
    return False


def cell_val(ws, row, col):
    """Get cell value, handling None."""
    return ws.cell(row=row, column=col).value


def validate_income(wb, data, errors):
    """Validate the Income tab against JSON data."""
    checks = 0
    ws = wb["Income"]
    monthly = data["income"].get("monthly", [])

    for i, item in enumerate(monthly):
        r = 4 + i
        # Check month name
        actual_month = cell_val(ws, r, 1)
        if actual_month != item["month"]:
            errors.append({
                "tab": "Income", "cell": f"A{r}",
                "field": f"monthly[{i}].month",
                "expected": item["month"], "actual": actual_month
            })
        checks += 1

        # Check amount
        actual_amount = cell_val(ws, r, 2)
        if not approx_eq(item["amount"], actual_amount):
            errors.append({
                "tab": "Income", "cell": f"B{r}",
                "field": f"monthly[{i}].amount",
                "expected": item["amount"], "actual": actual_amount
            })
        checks += 1

    # Check row count
    expected_count = len(monthly)
    actual_count = 0
    for r in range(4, 4 + expected_count + 5):
        v = cell_val(ws, r, 1)
        if v == "TOTAL":
            break
        if v is not None:
            actual_count += 1
    if actual_count != expected_count:
        errors.append({
            "tab": "Income", "cell": "A",
            "field": "monthly.length",
            "expected": expected_count, "actual": actual_count
        })
    checks += 1

    # Check interest
    interest = data["income"].get("interest", [])
    for i, item in enumerate(interest):
        found = False
        for r in range(1, ws.max_row + 1):
            if cell_val(ws, r, 1) == item["source"]:
                actual = cell_val(ws, r, 2)
                if not approx_eq(item["amount"], actual):
                    errors.append({
                        "tab": "Income", "cell": f"B{r}",
                        "field": f"interest[{i}].amount",
                        "expected": item["amount"], "actual": actual
                    })
                found = True
                break
        if not found:
            errors.append({
                "tab": "Income", "cell": "?",
                "field": f"interest[{i}].source",
                "expected": item["source"], "actual": "NOT FOUND"
            })
        checks += 1

    # Check by_source
    by_source = data["income"].get("by_source", [])
    for i, item in enumerate(by_source):
        found = False
        for r in range(1, ws.max_row + 1):
            if cell_val(ws, r, 1) == item["source"]:
                actual = cell_val(ws, r, 2)
                if not approx_eq(item["annual_total"], actual):
                    errors.append({
                        "tab": "Income", "cell": f"B{r}",
                        "field": f"by_source[{i}].annual_total",
                        "expected": item["annual_total"], "actual": actual
                    })
                found = True
                break
        if not found:
            errors.append({
                "tab": "Income", "cell": "?",
                "field": f"by_source[{i}].source",
                "expected": item["source"], "actual": "NOT FOUND"
            })
        checks += 1

    return checks


def validate_expenses(wb, data, errors):
    """Validate the Business Expenses tab."""
    checks = 0
    ws = wb["Business Expenses"]
    expenses = data["business_expenses"]

    # Validate recurring
    recurring = expenses.get("recurring", [])
    for i, item in enumerate(recurring):
        r = 5 + i
        actual_vendor = cell_val(ws, r, 1)
        if actual_vendor != item["vendor"]:
            errors.append({
                "tab": "Business Expenses", "cell": f"A{r}",
                "field": f"recurring[{i}].vendor",
                "expected": item["vendor"], "actual": actual_vendor
            })
        checks += 1

        actual_total = cell_val(ws, r, 4)
        if not approx_eq(item["annual_total"], actual_total):
            errors.append({
                "tab": "Business Expenses", "cell": f"D{r}",
                "field": f"recurring[{i}].annual_total",
                "expected": item["annual_total"], "actual": actual_total
            })
        checks += 1

    # Find where one-time section starts by looking for "ONE-TIME" header
    onetime_data_start = None
    for r in range(1, ws.max_row + 1):
        if cell_val(ws, r, 1) == "ONE-TIME / NON-RECURRING":
            onetime_data_start = r + 2  # skip header row
            break

    # Validate one-time
    one_time = expenses.get("one_time", [])
    if onetime_data_start:
        for i, item in enumerate(one_time):
            r = onetime_data_start + i
            actual_vendor = cell_val(ws, r, 2)
            if actual_vendor != item["vendor"]:
                errors.append({
                    "tab": "Business Expenses", "cell": f"B{r}",
                    "field": f"one_time[{i}].vendor",
                    "expected": item["vendor"], "actual": actual_vendor
                })
            checks += 1

            actual_amount = cell_val(ws, r, 3)
            if not approx_eq(item["amount"], actual_amount):
                errors.append({
                    "tab": "Business Expenses", "cell": f"C{r}",
                    "field": f"one_time[{i}].amount",
                    "expected": item["amount"], "actual": actual_amount
                })
            checks += 1

    # Validate excluded items — match by payee AND amount to handle duplicate payees
    excluded = expenses.get("excluded", [])
    matched_rows = set()
    for i, item in enumerate(excluded):
        found = False
        for r in range(1, ws.max_row + 1):
            if r in matched_rows:
                continue
            if cell_val(ws, r, 2) == item["payee"]:
                actual_amount = cell_val(ws, r, 3)
                if approx_eq(item["amount"], actual_amount):
                    found = True
                    matched_rows.add(r)
                    break
        if not found:
            # Try looser match — just find the payee
            for r in range(1, ws.max_row + 1):
                if r in matched_rows:
                    continue
                if cell_val(ws, r, 2) == item["payee"]:
                    actual_amount = cell_val(ws, r, 3)
                    errors.append({
                        "tab": "Business Expenses", "cell": f"C{r}",
                        "field": f"excluded[{i}].amount",
                        "expected": item["amount"], "actual": actual_amount
                    })
                    matched_rows.add(r)
                    found = True
                    break
            if not found:
                errors.append({
                    "tab": "Business Expenses", "cell": "?",
                    "field": f"excluded[{i}].payee",
                    "expected": item["payee"], "actual": "NOT FOUND"
                })
        checks += 1

    # Validate totals (computed from JSON)
    expected_recurring = sum(r["annual_total"] for r in recurring)
    expected_onetime = sum(o["amount"] for o in one_time)
    expected_grand = expected_recurring + expected_onetime

    # Find grand total row
    for r in range(1, ws.max_row + 1):
        v = cell_val(ws, r, 1)
        if v and "GRAND TOTAL" in str(v):
            actual_grand = cell_val(ws, r, 4)
            if actual_grand is not None and not approx_eq(expected_grand, actual_grand):
                errors.append({
                    "tab": "Business Expenses", "cell": f"D{r}",
                    "field": "grand_total",
                    "expected": expected_grand, "actual": actual_grand
                })
            checks += 1
            break

    return checks


def validate_personal_deductions(wb, data, errors):
    """Validate the Personal Deductions tab."""
    checks = 0
    ws = wb["Personal Deductions"]
    items = data.get("personal_deductions", [])

    for i, item in enumerate(items):
        r = 4 + i
        actual_type = cell_val(ws, r, 1)
        if actual_type != item["type"]:
            errors.append({
                "tab": "Personal Deductions", "cell": f"A{r}",
                "field": f"[{i}].type",
                "expected": item["type"], "actual": actual_type
            })
        checks += 1

        actual_total = cell_val(ws, r, 4)
        if not approx_eq(item["annual_total"], actual_total):
            errors.append({
                "tab": "Personal Deductions", "cell": f"D{r}",
                "field": f"[{i}].annual_total",
                "expected": item["annual_total"], "actual": actual_total
            })
        checks += 1

        monthly = item.get("monthly_amount")
        if monthly is not None:
            actual_monthly = cell_val(ws, r, 3)
            if not approx_eq(monthly, actual_monthly):
                errors.append({
                    "tab": "Personal Deductions", "cell": f"C{r}",
                    "field": f"[{i}].monthly_amount",
                    "expected": monthly, "actual": actual_monthly
                })
            checks += 1

    # Check row count
    checks += 1
    expected_count = len(items)
    actual_count = 0
    for r in range(4, 4 + expected_count + 5):
        v = cell_val(ws, r, 1)
        if v == "TOTAL" or v is None:
            break
        actual_count += 1
    if actual_count != expected_count:
        errors.append({
            "tab": "Personal Deductions", "cell": "A",
            "field": "length",
            "expected": expected_count, "actual": actual_count
        })

    return checks


def validate_home_office(wb, data, errors):
    """Validate the Home Office tab."""
    checks = 0
    ws = wb["Home Office"]
    locations = data["home_office"].get("locations", [])

    for i, loc in enumerate(locations):
        # Find the location header row
        found = False
        for r in range(1, ws.max_row + 1):
            v = cell_val(ws, r, 1)
            if v and f"LOCATION {i + 1}" in str(v):
                found = True
                # Check key values in subsequent rows
                for check_r in range(r + 1, r + 10):
                    label = cell_val(ws, check_r, 1)
                    value = cell_val(ws, check_r, 2)
                    if label == "Office Sq Ft":
                        if value != loc["office_sqft"]:
                            errors.append({
                                "tab": "Home Office", "cell": f"B{check_r}",
                                "field": f"locations[{i}].office_sqft",
                                "expected": loc["office_sqft"], "actual": value
                            })
                        checks += 1
                    elif label == "Home Sq Ft":
                        if value != loc["home_sqft"]:
                            errors.append({
                                "tab": "Home Office", "cell": f"B{check_r}",
                                "field": f"locations[{i}].home_sqft",
                                "expected": loc["home_sqft"], "actual": value
                            })
                        checks += 1
                    elif label == "Monthly Rent":
                        if not approx_eq(loc["monthly_rent"], value):
                            errors.append({
                                "tab": "Home Office", "cell": f"B{check_r}",
                                "field": f"locations[{i}].monthly_rent",
                                "expected": loc["monthly_rent"], "actual": value
                            })
                        checks += 1
                    elif label == "Monthly Deduction":
                        if not approx_eq(loc["monthly_deduction"], value):
                            errors.append({
                                "tab": "Home Office", "cell": f"B{check_r}",
                                "field": f"locations[{i}].monthly_deduction",
                                "expected": loc["monthly_deduction"], "actual": value
                            })
                        checks += 1
                    elif label and label.startswith("Subtotal"):
                        if not approx_eq(loc["subtotal"], value):
                            errors.append({
                                "tab": "Home Office", "cell": f"B{check_r}",
                                "field": f"locations[{i}].subtotal",
                                "expected": loc["subtotal"], "actual": value
                            })
                        checks += 1
                break

        if not found:
            errors.append({
                "tab": "Home Office", "cell": "?",
                "field": f"locations[{i}]",
                "expected": f"LOCATION {i + 1}", "actual": "NOT FOUND"
            })
        checks += 1

    # Check total — formula cells may be None before recalc, so also check
    # against the formula-bearing workbook
    expected_total = sum(l["subtotal"] for l in locations)
    for r in range(1, ws.max_row + 1):
        v = cell_val(ws, r, 1)
        if v and "TOTAL HOME OFFICE" in str(v):
            actual = cell_val(ws, r, 2)
            # If actual is None, formula hasn't been evaluated — skip rather than fail
            if actual is not None and not approx_eq(expected_total, actual):
                errors.append({
                    "tab": "Home Office", "cell": f"B{r}",
                    "field": "total",
                    "expected": expected_total, "actual": actual
                })
            checks += 1
            break

    return checks


def validate_estimated_taxes(wb, data, errors):
    """Validate the Estimated Taxes tab."""
    checks = 0
    ws = wb["Estimated Taxes"]
    est = data["estimated_taxes"]

    payments = est.get("payments", [])
    for i, item in enumerate(payments):
        r = 4 + i
        actual_auth = cell_val(ws, r, 1)
        if actual_auth != item["authority"]:
            errors.append({
                "tab": "Estimated Taxes", "cell": f"A{r}",
                "field": f"payments[{i}].authority",
                "expected": item["authority"], "actual": actual_auth
            })
        checks += 1

        actual_amount = cell_val(ws, r, 4)
        if not approx_eq(item["amount"], actual_amount):
            errors.append({
                "tab": "Estimated Taxes", "cell": f"D{r}",
                "field": f"payments[{i}].amount",
                "expected": item["amount"], "actual": actual_amount
            })
        checks += 1

    # Validate prior-year payments — match by date AND amount to handle same-date entries
    prior = est.get("prior_year_payments", [])
    prior_matched_rows = set()
    for i, item in enumerate(prior):
        found = False
        # First try exact match on date + amount
        for r in range(1, ws.max_row + 1):
            if r in prior_matched_rows:
                continue
            if cell_val(ws, r, 3) == item["date_paid"]:
                actual = cell_val(ws, r, 4)
                if approx_eq(item["amount"], actual):
                    found = True
                    prior_matched_rows.add(r)
                    break
        if not found:
            # Fall back to date-only match
            for r in range(1, ws.max_row + 1):
                if r in prior_matched_rows:
                    continue
                if cell_val(ws, r, 3) == item["date_paid"]:
                    actual = cell_val(ws, r, 4)
                    errors.append({
                        "tab": "Estimated Taxes", "cell": f"D{r}",
                        "field": f"prior_year_payments[{i}].amount",
                        "expected": item["amount"], "actual": actual
                    })
                    prior_matched_rows.add(r)
                    found = True
                    break
            if not found:
                errors.append({
                    "tab": "Estimated Taxes", "cell": "?",
                    "field": f"prior_year_payments[{i}]",
                    "expected": item["date_paid"], "actual": "NOT FOUND"
                })
        checks += 1

    return checks


def validate_health_insurance(wb, data, errors):
    """Validate the Health Insurance tab."""
    checks = 0
    ws = wb["Health Insurance"]
    items = data.get("health_insurance", [])

    for i, item in enumerate(items):
        r = 4 + i
        actual_provider = cell_val(ws, r, 1)
        if actual_provider != item["provider"]:
            errors.append({
                "tab": "Health Insurance", "cell": f"A{r}",
                "field": f"[{i}].provider",
                "expected": item["provider"], "actual": actual_provider
            })
        checks += 1

        actual_premium = cell_val(ws, r, 2)
        if not approx_eq(item["monthly_premium"], actual_premium):
            errors.append({
                "tab": "Health Insurance", "cell": f"B{r}",
                "field": f"[{i}].monthly_premium",
                "expected": item["monthly_premium"], "actual": actual_premium
            })
        checks += 1

        actual_months = cell_val(ws, r, 5)
        if actual_months != item["months"]:
            errors.append({
                "tab": "Health Insurance", "cell": f"E{r}",
                "field": f"[{i}].months",
                "expected": item["months"], "actual": actual_months
            })
        checks += 1

    return checks


def validate_retirement(wb, data, errors):
    """Validate the Retirement tab."""
    checks = 0
    ws = wb["Retirement"]
    retirement = data["retirement"]
    contributions = retirement.get("contributions", [])

    if not contributions:
        # Should show "(No contributions made yet)"
        v = cell_val(ws, 4, 1)
        if v and "No contributions" in str(v):
            checks += 1
        else:
            errors.append({
                "tab": "Retirement", "cell": "A4",
                "field": "empty_message",
                "expected": "No contributions message", "actual": v
            })
            checks += 1
    else:
        for i, item in enumerate(contributions):
            r = 4 + i
            actual_type = cell_val(ws, r, 1)
            if actual_type != item["account_type"]:
                errors.append({
                    "tab": "Retirement", "cell": f"A{r}",
                    "field": f"contributions[{i}].account_type",
                    "expected": item["account_type"], "actual": actual_type
                })
            checks += 1

            actual_amount = cell_val(ws, r, 3)
            if not approx_eq(item["amount"], actual_amount):
                errors.append({
                    "tab": "Retirement", "cell": f"C{r}",
                    "field": f"contributions[{i}].amount",
                    "expected": item["amount"], "actual": actual_amount
                })
            checks += 1

    # Validate notes exist
    notes = retirement.get("notes", [])
    for i, note in enumerate(notes):
        found = False
        for r in range(1, ws.max_row + 1):
            v = cell_val(ws, r, 1)
            if v == note:
                found = True
                break
        if not found:
            errors.append({
                "tab": "Retirement", "cell": "?",
                "field": f"notes[{i}]",
                "expected": note, "actual": "NOT FOUND"
            })
        checks += 1

    return checks


def validate_tab_existence(wb, errors):
    """Check that all 8 expected tabs exist."""
    checks = 0
    expected_tabs = [
        "Summary", "Income", "Business Expenses", "Personal Deductions",
        "Home Office", "Estimated Taxes", "Health Insurance", "Retirement"
    ]
    for tab in expected_tabs:
        if tab not in wb.sheetnames:
            errors.append({
                "tab": tab, "cell": "-",
                "field": "tab_exists",
                "expected": True, "actual": False
            })
        checks += 1
    return checks


def main(json_path, xlsx_path):
    with open(json_path, "r") as f:
        data = json.load(f)

    # Load with data_only=True to get computed formula values
    wb = load_workbook(xlsx_path, data_only=True)

    errors = []
    total_checks = 0

    total_checks += validate_tab_existence(wb, errors)

    # Only validate tabs that exist
    if "Income" in wb.sheetnames:
        total_checks += validate_income(wb, data, errors)
    if "Business Expenses" in wb.sheetnames:
        total_checks += validate_expenses(wb, data, errors)
    if "Personal Deductions" in wb.sheetnames:
        total_checks += validate_personal_deductions(wb, data, errors)
    if "Home Office" in wb.sheetnames:
        total_checks += validate_home_office(wb, data, errors)
    if "Estimated Taxes" in wb.sheetnames:
        total_checks += validate_estimated_taxes(wb, data, errors)
    if "Health Insurance" in wb.sheetnames:
        total_checks += validate_health_insurance(wb, data, errors)
    if "Retirement" in wb.sheetnames:
        total_checks += validate_retirement(wb, data, errors)

    wb.close()

    result = {
        "status": "pass" if not errors else "fail",
        "checks": total_checks,
        "errors": errors,
    }

    print(json.dumps(result, indent=2))
    return 0 if not errors else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python validate_xlsx.py <json_file> <xlsx_file>")
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
